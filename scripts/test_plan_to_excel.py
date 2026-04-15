#!/usr/bin/env python3
"""
Export TEST_PLAN.md to Excel (.xlsx) with formatted sheets per category.

Usage:
    pip install openpyxl
    python scripts/test_plan_to_excel.py [-i TEST_PLAN.md] [-o TEST_PLAN.xlsx]

The script parses markdown tables grouped by level-2 headings (## Section name)
and creates one sheet per section. Cells are formatted for readable tracking.

Example workflow:
    1. Run this to produce TEST_PLAN.xlsx
    2. Share the .xlsx with the team for execution
    3. Testers fill Status/Actual Result/Tested By/Tested At columns
    4. Re-import is not automatic - copy results back to TEST_PLAN.md manually
       or use the companion tool (future work)
"""

import argparse
import re
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


SHEET_NAME_MAX_LENGTH = 31

STATUS_COLORS = {
    "pass": "C6EFCE",
    "fail": "FFC7CE",
    "partial": "FFEB9C",
    "skipped": "D9D9D9",
    "blocked": "F2DCDB",
}

PRIORITY_COLORS = {
    "P0": "FF6B6B",
    "P1": "FFD93D",
    "P2": "6BCB77",
    "P3": "4D96FF",
}


def sanitize_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ]"""
    name = re.sub(r'[:\\/?*\[\]]', '-', name)
    if len(name) > SHEET_NAME_MAX_LENGTH:
        name = name[:SHEET_NAME_MAX_LENGTH - 3] + "..."
    return name.strip() or "Sheet"


def parse_sections(text: str) -> list[tuple[str, list[dict]]]:
    """Parse level-2 sections with tables. Returns [(section_title, rows), ...]."""
    # Split by ## headings
    lines = text.split("\n")
    sections: list[tuple[str, list[str]]] = []
    current_title: str | None = None
    current_lines: list[str] = []

    for line in lines:
        h2_match = re.match(r"^##\s+(.+?)\s*$", line)
        if h2_match:
            if current_title is not None:
                sections.append((current_title, current_lines))
            current_title = h2_match.group(1).strip()
            current_lines = []
        elif current_title is not None:
            current_lines.append(line)

    if current_title is not None:
        sections.append((current_title, current_lines))

    # For each section, extract the first markdown table and its rows
    parsed: list[tuple[str, list[dict]]] = []
    for title, body in sections:
        rows = extract_first_table(body)
        if rows:
            parsed.append((title, rows))

    return parsed


def extract_first_table(body_lines: list[str]) -> list[dict]:
    """Find the first markdown table and return a list of row dicts."""
    headers: list[str] | None = None
    rows: list[dict] = []
    in_table = False

    for line in body_lines:
        stripped = line.strip()
        if not stripped:
            if in_table:
                # empty line ends the table
                break
            continue

        if stripped.startswith("|") and stripped.endswith("|"):
            cells = [c.strip() for c in stripped.strip("|").split("|")]
            if headers is None:
                headers = cells
                in_table = True
                continue
            # Skip the --- separator row
            if all(set(c.replace(":", "").replace("-", "").strip()) <= {""} for c in cells):
                continue
            # Normalize row length
            while len(cells) < len(headers):
                cells.append("")
            rows.append({h: v for h, v in zip(headers, cells)})
        else:
            if in_table:
                break

    return rows


def apply_header_style(cell):
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill("solid", fgColor="2F5597")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="888888"),
        right=Side(style="thin", color="888888"),
        top=Side(style="thin", color="888888"),
        bottom=Side(style="thin", color="888888"),
    )


def apply_cell_style(cell, wrap=True):
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )


def colorize_priority(cell):
    val = (cell.value or "").upper().strip()
    color = PRIORITY_COLORS.get(val)
    if color:
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def colorize_status(cell):
    val = (cell.value or "").lower().strip()
    # Also handle unicode status symbols
    mapping = {
        "\u2705": "pass", "pass": "pass", "ok": "pass",
        "\u274c": "fail", "fail": "fail",
        "\u26a0": "partial", "partial": "partial",
        "\u23f8": "skipped", "skipped": "skipped", "skip": "skipped",
        "\u2753": "blocked", "blocked": "blocked",
    }
    key = mapping.get(val)
    if key:
        cell.fill = PatternFill("solid", fgColor=STATUS_COLORS[key])
        cell.alignment = Alignment(horizontal="center", vertical="center")


def set_column_widths(ws, headers: list[str]):
    preferred = {
        "id": 14, "priority": 9, "description": 38,
        "steps": 45, "expected result": 45, "status": 12,
        "tested by": 14, "tested at": 14, "notes": 30,
        "platform": 12, "version": 10,
    }
    for i, h in enumerate(headers, start=1):
        key = h.lower().strip()
        width = preferred.get(key, 20)
        ws.column_dimensions[get_column_letter(i)].width = width


def build_workbook(sections: list[tuple[str, list[dict]]], source_file: str) -> Workbook:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Index / summary sheet
    index_ws = wb.create_sheet("Index")
    index_ws.append(["Section", "Test Count", "Sheet"])
    for c in index_ws[1]:
        apply_header_style(c)

    for title, rows in sections:
        sheet_name = sanitize_sheet_name(title)
        # Avoid collisions
        base = sheet_name
        i = 1
        while sheet_name in wb.sheetnames:
            suffix = f"_{i}"
            sheet_name = (base[:SHEET_NAME_MAX_LENGTH - len(suffix)] + suffix)
            i += 1

        ws = wb.create_sheet(sheet_name)
        headers = list(rows[0].keys())
        # Add extended tracking columns if missing
        for extra in ("Tested By", "Tested At", "Notes"):
            if extra not in headers:
                headers.append(extra)

        ws.append(headers)
        for cell in ws[1]:
            apply_header_style(cell)

        for row_dict in rows:
            row = [row_dict.get(h, "") for h in headers]
            ws.append(row)

        # Style body rows
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                apply_cell_style(cell)
                header_value = ws.cell(row=1, column=col_idx).value or ""
                h_lower = header_value.lower()
                if "priority" in h_lower:
                    colorize_priority(cell)
                elif h_lower == "status":
                    colorize_status(cell)

        set_column_widths(ws, headers)
        ws.freeze_panes = "A2"

        # Add summary to index
        index_ws.append([title, len(rows), sheet_name])

    # Style index sheet
    for row_idx in range(2, index_ws.max_row + 1):
        for col_idx in range(1, index_ws.max_column + 1):
            apply_cell_style(index_ws.cell(row=row_idx, column=col_idx))
    index_ws.column_dimensions["A"].width = 40
    index_ws.column_dimensions["B"].width = 12
    index_ws.column_dimensions["C"].width = 35
    index_ws.freeze_panes = "A2"

    # Metadata sheet
    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["Generated from", source_file])
    meta_ws.append(["Total sections", len(sections)])
    meta_ws.append(["Total tests", sum(len(r) for _, r in sections)])
    for row in meta_ws.iter_rows():
        for cell in row:
            apply_cell_style(cell)
    meta_ws.column_dimensions["A"].width = 20
    meta_ws.column_dimensions["B"].width = 50

    return wb


def main():
    parser = argparse.ArgumentParser(description="Export TEST_PLAN.md to Excel .xlsx")
    parser.add_argument("-i", "--input", default="TEST_PLAN.md", help="Input markdown file")
    parser.add_argument("-o", "--output", default="TEST_PLAN.xlsx", help="Output Excel file")
    args = parser.parse_args()

    path = Path(args.input)
    if not path.exists():
        print(f"ERROR: input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    text = path.read_text(encoding="utf-8")
    sections = parse_sections(text)

    if not sections:
        print("ERROR: no sections with tables found in markdown", file=sys.stderr)
        sys.exit(1)

    wb = build_workbook(sections, str(path))
    wb.save(args.output)

    total = sum(len(r) for _, r in sections)
    print(f"[OK] Wrote {args.output}")
    print(f"  Sections: {len(sections)}")
    print(f"  Total tests: {total}")


if __name__ == "__main__":
    main()
